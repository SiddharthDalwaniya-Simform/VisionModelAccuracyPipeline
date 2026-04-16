"""
excel_report.py — Creates and updates the test results spreadsheet.

Writes one row per video with color-coded status.
Green = PASS, Red = NO EVENT, Yellow = MISMATCH.
"""


import logging 
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config

log = logging.getLogger("excel")

# Colors
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
ORANGE = PatternFill("solid", fgColor="FF8C00")  # connection failures
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF", name="Arial")
DATA_FONT = Font(name="Arial", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

HEADERS = [
    "Sr. No.",
    "Offline Video Name",
    "Video Duration (s)",
    "Event Received?",
    "S3 Event Clip Link",
    "Match Result",
    "Match Ratio",
    "Status",
    "Timestamp",
    "Notes",
]

COL_WIDTHS = [8, 35, 18, 16, 50, 14, 14, 12, 22, 40]


class ExcelReport:

    def __init__(self):
        self.path = config.OUTPUT_EXCEL
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
            # Find the next empty row so we append instead of overwrite
            self._start_row = self._find_next_empty_row()
            log.info("Excel file exists — will append starting at row %d",
                     self._start_row)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Test Results"
            self._write_header()
            self._start_row = 2  # row 1 is the header

    def _find_next_empty_row(self) -> int:
        """Find the first empty row after the header."""
        row = 2
        while self.ws.cell(row=row, column=1).value is not None:
            row += 1
        return row

    def _write_header(self):
        for col, header in enumerate(HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        for i, w in enumerate(COL_WIDTHS, 1):
            col_letter = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
            self.ws.column_dimensions[col_letter].width = w

        self.ws.auto_filter.ref = "A1:J1"
        self.ws.freeze_panes = "A2"

    def add_result(
        self,
        sr_no: int,
        video_name: str,
        duration: float,
        event_received: bool,
        s3_link: str,
        match_result,
        match_ratio: float,
        notes: str = "",
        force_status: str = None,
    ):
        row = self._start_row + (sr_no - 1)

        if force_status:
            status, status_fill = force_status, ORANGE
        elif event_received and match_result:
            status, status_fill = "PASS", GREEN
        elif event_received and not match_result:
            status, status_fill = "MISMATCH", YELLOW
        else:
            status, status_fill = "NO EVENT", RED

        values = [
            sr_no,
            video_name,
            round(duration, 1),
            "YES" if event_received else "NO",
            s3_link or "—",
            "MATCH" if match_result else ("NO MATCH" if match_result is False else "N/A"),
            round(match_ratio, 4) if match_ratio > 0 else "—",
            status,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            notes,
        ]

        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER
            cell.border = BORDER

        self.ws.cell(row=row, column=8).fill = status_fill
        self.ws.cell(row=row, column=4).fill = GREEN if event_received else RED

        self.wb.save(self.path)
        log.info("  Excel row %d → %s", row, status)

    def add_summary(self, total: int, passed: int, missed: int, mismatched: int, execution_time: str = ""):
        last_data_row = self._start_row + total - 1
        row = last_data_row + 2
        bold = Font(bold=True, name="Arial", size=12)

        self.ws.cell(row=row, column=1, value="SUMMARY").font = bold
        self.ws.cell(row=row + 1, column=1, value="Total Videos")
        self.ws.cell(row=row + 1, column=2, value=total)
        self.ws.cell(row=row + 2, column=1, value="Events Detected (PASS)")
        self.ws.cell(row=row + 2, column=2, value=passed)
        self.ws.cell(row=row + 2, column=2).fill = GREEN
        self.ws.cell(row=row + 3, column=1, value="Missed (NO EVENT)")
        self.ws.cell(row=row + 3, column=2, value=missed)
        self.ws.cell(row=row + 3, column=2).fill = RED
        self.ws.cell(row=row + 4, column=1, value="Mismatched")
        self.ws.cell(row=row + 4, column=2, value=mismatched)
        self.ws.cell(row=row + 4, column=2).fill = YELLOW
        self.ws.cell(row=row + 5, column=1, value="Detection Rate")
        self.ws.cell(row=row + 5, column=2, value=f"=B{row+2}/B{row+1}*100")
        self.ws.cell(row=row + 5, column=2).number_format = '0.0"%"'
        self.ws.cell(row=row + 6, column=1, value="Total Execution Time")
        self.ws.cell(row=row + 6, column=2, value=execution_time)

        self.wb.save(self.path)
        log.info("Summary added to Excel.")
